"""
from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
# ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample12.xlsx")
"""


"""
# create xl file
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws1 = wb.create_sheet('sheet1')
ws2 = wb.create_sheet('sheet2')
ws3 = wb.create_sheet('sheet3')

wb.save("Myxl.xlsx")
"""


'''
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws1 = wb.create_sheet('sheet1')
ws2 = wb.create_sheet('sheet2')
ws3 = wb.create_sheet('sheet3')

columns = list(range(0,10))

for col in ws.iter_cols(min_col=3, max_col=5,max_row=5):
    for cell in col:
        ws1.append(columns)

wb.save("Myxl.xlsx")
'''



from openpyxl import Workbook

wb = Workbook()

ws = wb.active

ws["A1"] = "Roll Number"
ws["B1"] = "Name"
ws["C1"] = "Mark"

wb.save("mark.xlsx")




